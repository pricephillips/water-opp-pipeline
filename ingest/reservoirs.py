from __future__ import annotations
"""
ingest/reservoirs.py
---------------------
Builds county-level infrastructure elasticity data from:
  1. Bureau of Reclamation (USBR) — reservoir storage levels for western US
     API: https://data.usbr.gov/api/
  2. Army Corps of Engineers National Inventory of Dams (NID)
     Download: https://nid.usace.army.mil (manual download required; see below)
  3. WRI Aqueduct 4.0 — seasonal variability component (manual download)

USBR covers ~20 major western reservoirs (Lake Mead, Hoover, Powell, Shasta, etc.)
NID covers ~90,000 dams nationally with capacity in acre-feet and county FIPS.

Manual downloads required (add to data/raw/):
  - NID: Download "NID_National.xlsx" from https://nid.usace.army.mil/api/nation/xlsx
    Save as: data/raw/NID_National.xlsx
  - WRI Aqueduct: Download "Aqueduct40_Y2023D07_baseline_monthly_gpkg.zip" from
    https://datasets.wri.org/dataset/aqueduct40
    Extract to: data/raw/wri_aqueduct/

Outputs: data/processed/reservoirs.csv
"""

import csv
import json
import os
import time
import urllib.request
import urllib.error
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
OUT_DIR = ROOT / "data" / "processed"
OUT_DIR.mkdir(parents=True, exist_ok=True)

USBR_API = "https://data.usbr.gov/rise/api"

# Major USBR reservoir → county FIPS mapping
# These are the primary storage reservoirs that serve metropolitan areas
# Keyed by USBR reservoir "parameter code" (catalog ID)
USBR_KEY_RESERVOIRS = {
    # Lake Mead (NV/AZ) → Clark County NV + Mohave County AZ
    "36": {"name": "Lake Mead", "fips": ["32003", "04015"], "capacity_af": 26_120_000},
    # Lake Powell (AZ/UT)
    "913": {"name": "Lake Powell", "fips": ["04013", "49037"], "capacity_af": 24_322_000},
    # Hoover Dam / Lake Mead (same reservoir, different USBR catalog entries)
    # Shasta Lake (CA) → Shasta County
    "1394": {"name": "Shasta Lake", "fips": ["06089"], "capacity_af": 4_552_000},
    # Oroville (CA) → Butte County
    "1392": {"name": "Lake Oroville", "fips": ["06007"], "capacity_af": 3_537_577},
    # Folsom (CA) → Sacramento / El Dorado
    "1388": {"name": "Folsom Lake", "fips": ["06067", "06017"], "capacity_af": 977_000},
    # Flaming Gorge (WY/UT)
    "920": {"name": "Flaming Gorge", "fips": ["56037", "49043"], "capacity_af": 3_788_700},
    # Elephant Butte (NM) → Sierra County
    "408": {"name": "Elephant Butte", "fips": ["35051"], "capacity_af": 2_065_625},
    # Horsetooth (CO) → Larimer County
    "332": {"name": "Horsetooth Reservoir", "fips": ["08069"], "capacity_af": 156_735},
    # Vallecito (CO)
    "342": {"name": "Vallecito", "fips": ["08067"], "capacity_af": 129_576},
    # Blue Mesa (CO) → Gunnison County
    "331": {"name": "Blue Mesa", "fips": ["08051"], "capacity_af": 940_800},
    # Glen Canyon (AZ)
    "914": {"name": "Glen Canyon", "fips": ["04013"], "capacity_af": 24_322_000},
}


def fetch_json(url: str, retries: int = 3) -> dict | list | None:
    headers = {"User-Agent": "water-risk-pipeline/1.0", "Accept": "application/json"}
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.loads(r.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            print(f"  HTTP {e.code} attempt {attempt+1}: {url[:80]}")
            if attempt < retries - 1:
                time.sleep(2)
        except Exception as e:
            print(f"  Error attempt {attempt+1}: {e}")
            if attempt < retries - 1:
                time.sleep(2)
    return None


def fetch_usbr_reservoir_storage(catalog_id: str) -> float | None:
    """
    Fetch current reservoir storage percentage from USBR RISE API.
    
    Returns current storage as percentage of total capacity (0–100),
    or None if unavailable.
    
    USBR RISE API: https://data.usbr.gov/rise/api/result/download?
        itemId={catalog_id}&type=json&optionType=day&beforeDt=...
    """
    url = (
        f"{USBR_API}/result/download"
        f"?itemId={catalog_id}"
        f"&type=json"
        f"&optionType=day"
    )
    data = fetch_json(url)
    if data is None or not data:
        return None

    # RISE API returns list of observations; take the most recent
    try:
        items = data if isinstance(data, list) else data.get("data", [])
        if not items:
            return None
        # Most recent entry has highest date
        most_recent = sorted(items, key=lambda x: x.get("dateTime", ""))[-1]
        storage_af = float(most_recent.get("result", 0) or 0)
        capacity = USBR_KEY_RESERVOIRS.get(catalog_id, {}).get("capacity_af", 1)
        pct = min(100.0, round((storage_af / capacity) * 100, 1))
        return pct
    except Exception as e:
        print(f"  USBR parse error for {catalog_id}: {e}")
        return None


def load_nid_data(nid_path: Path) -> dict[str, dict]:
    """
    Load Army Corps NID data and aggregate reservoir capacity to county FIPS.

    NID Excel has columns including:
        FederalId, DamName, NidId, State, CountyName, Latitude, Longitude,
        PrimaryPurpose, NormalStorageAcFt, MaxStorageAcFt, SurfaceAreaAcres,
        CountyFips (added in recent versions), or state + county for FIPS lookup

    Returns: {fips: {'total_reservoir_capacity_af': float, 'dam_count': int}}
    """
    if not nid_path.exists():
        print(f"  NID file not found at {nid_path}")
        print("  Download from: https://nid.usace.army.mil/api/nation/xlsx")
        print("  Save as: data/raw/NID_National.xlsx")
        return {}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(nid_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Map column names to indices
        col = {name: idx for idx, name in enumerate(headers)}

        required = ["NormalStorageAcFt", "State"]
        missing = [c for c in required if c not in col]
        if missing:
            print(f"  NID column mismatch. Expected {required}, got {headers[:10]}")
            return {}

        county_data = defaultdict(lambda: {"total_reservoir_capacity_af": 0.0, "dam_count": 0})

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Get FIPS if available
            fips = ""
            if "CountyFips" in col:
                raw_fips = str(row[col["CountyFips"]] or "").strip()
                fips = raw_fips.zfill(5) if raw_fips else ""

            if not fips:
                continue  # Skip if no FIPS mapping

            # Normal storage (acre-feet)
            storage_raw = row[col["NormalStorageAcFt"]]
            try:
                storage = float(storage_raw or 0)
            except (ValueError, TypeError):
                storage = 0.0

            if storage <= 0:
                continue

            county_data[fips]["total_reservoir_capacity_af"] += storage
            county_data[fips]["dam_count"] += 1

        print(f"  NID: loaded {sum(d['dam_count'] for d in county_data.values())} dams across {len(county_data)} counties")
        return dict(county_data)

    except ImportError:
        print("  openpyxl not installed. Run: pip install openpyxl")
        return {}
    except Exception as e:
        print(f"  NID load error: {e}")
        return {}


def load_census_population(pop_path: Path | None = None) -> dict[str, int]:
    """
    Load county population estimates for per-capita reservoir calculation.
    
    Falls back to a built-in lookup for major counties if file not available.
    Full dataset: https://www.census.gov/data/tables/time-series/demo/popest/2020s-counties-total.html
    """
    if pop_path and pop_path.exists():
        pop = {}
        with open(pop_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                fips = row.get("fips", row.get("FIPS", "")).zfill(5)
                pop_val = row.get("population", row.get("POPESTIMATE2023", 0))
                try:
                    pop[fips] = int(float(pop_val or 0))
                except (ValueError, TypeError):
                    pass
        return pop

    # Built-in fallback for 50 largest counties by population (2023 estimates)
    return {
        "06037": 9_829_544,   # Los Angeles
        "17031": 5_150_233,   # Cook (Chicago)
        "48201": 4_780_913,   # Harris (Houston)
        "04013": 4_585_871,   # Maricopa (Phoenix)
        "06073": 3_322_287,   # San Diego
        "48113": 2_785_984,   # Dallas
        "62000": 2_673_149,   # NYC (approx individual boroughs separately)
        "06059": 3_222_498,   # Orange CA
        "12086": 2_688_162,   # Miami-Dade
        "48439": 2_336_853,   # Tarrant (Fort Worth)
        "53033": 2_269_675,   # King (Seattle)
        "06065": 2_437_864,   # Riverside
        "36047": 2_576_771,   # Kings (Brooklyn)
        "48029": 2_006_954,   # Bexar (San Antonio)
        "36081": 2_278_906,   # Queens
        "06071": 2_175_188,   # San Bernardino
        "06085": 1_894_215,   # Santa Clara
        "12057": 1_420_098,   # Hillsborough (Tampa)
        "32003": 2_283_929,   # Clark (Las Vegas)
        "48453": 1_300_000,   # Travis (Austin)
        "13121": 1_100_000,   # Fulton (Atlanta)
        "51059": 1_150_000,   # Fairfax (Northern VA)
        "39035": 1_256_894,   # Cuyahoga (Cleveland)
        "25025":   798_552,   # Suffolk (Boston)
        "37119":   1_115_263, # Mecklenburg (Charlotte)
        "55079":   950_395,   # Milwaukee
        "39049": 1_323_807,   # Franklin (Columbus)
        "26163": 1_739_763,   # Wayne (Detroit)
        "08031":   715_522,   # Denver
        "49035": 1_175_000,   # Salt Lake
        "12095":   754_500,   # Orange (Orlando)
        "06001": 1_666_753,   # Alameda (Oakland)
        "53061":   835_622,   # Snohomish (Everett)
        "42101": 1_603_797,   # Philadelphia
        "40109":   804_281,   # Oklahoma
        "22071": 1_000_000,   # Orleans (New Orleans)
        "37183":   1_062_880, # Wake (Raleigh)
        "12031":   995_567,   # Duval (Jacksonville)
        "29189":   1_010_000, # St. Louis
        "28049":   404_132,   # Hinds (Jackson MS)
        "45045":   527_234,   # Greenville SC
        "26081": 1_392_000,   # Kent (Grand Rapids)
        "19153":   509_000,   # Polk (Des Moines)
        "48085": 2_636_905,   # Collin (Plano TX)
        "36005": 1_427_056,   # Bronx
        "48121": 1_100_000,   # Denton (TX)
        "34013": 2_000_000,   # Essex (Newark NJ area)
    }


def compute_reservoir_scores(
    nid_data: dict[str, dict],
    usbr_current: dict[str, float],
    population: dict[str, int],
) -> dict[str, dict]:
    """
    Combine NID capacity, USBR current storage, and population into
    per-county elasticity metrics.

    Returns: {fips: {reservoir_capacity_af, reservoir_capacity_per_capita,
                     reservoir_score, usbr_storage_pct, reservoir_current_score}}
    """
    all_fips = set(nid_data.keys()) | set(usbr_current.keys())
    results = {}

    # Find max per-capita for normalization
    per_capitas = []
    for fips in all_fips:
        cap = nid_data.get(fips, {}).get("total_reservoir_capacity_af", 0)
        pop = population.get(fips, 100_000)
        if pop > 0 and cap > 0:
            per_capitas.append(cap / pop)

    max_per_capita = max(per_capitas) if per_capitas else 1.0
    # Cap at 95th percentile to avoid outliers (tiny counties with huge dams)
    per_capitas_sorted = sorted(per_capitas)
    p95_idx = int(len(per_capitas_sorted) * 0.95)
    p95 = per_capitas_sorted[p95_idx] if p95_idx < len(per_capitas_sorted) else max_per_capita

    for fips in all_fips:
        cap = nid_data.get(fips, {}).get("total_reservoir_capacity_af", 0)
        dam_ct = nid_data.get(fips, {}).get("dam_count", 0)
        pop = population.get(fips, 50_000)
        per_capita = (cap / pop) if pop > 0 and cap > 0 else 0.0
        # Normalize to 0-100 capped at p95
        reservoir_score = min(100.0, round((per_capita / p95) * 100, 1)) if p95 > 0 else 0.0

        storage_pct = usbr_current.get(fips)
        current_score = round(100 - storage_pct, 1) if storage_pct is not None else None

        results[fips] = {
            "reservoir_capacity_af": round(cap),
            "dam_count": dam_ct,
            "reservoir_capacity_per_capita": round(per_capita, 2),
            "reservoir_score": reservoir_score,
            "usbr_storage_pct": storage_pct,
            "reservoir_current_score": current_score,
        }

    return results


def run():
    print("=== Reservoir & Infrastructure Ingestion ===")

    # 1. USBR current storage for key western reservoirs
    print("Fetching USBR reservoir storage ...")
    usbr_current: dict[str, float] = {}
    for catalog_id, meta in USBR_KEY_RESERVOIRS.items():
        pct = fetch_usbr_reservoir_storage(catalog_id)
        if pct is not None:
            print(f"  {meta['name']}: {pct:.1f}% of capacity")
            for fips in meta["fips"]:
                # Multiple reservoirs can serve the same county; take minimum (worst case)
                if fips not in usbr_current or pct < usbr_current[fips]:
                    usbr_current[fips] = pct
        time.sleep(0.5)

    # 2. NID dam capacity data
    nid_path = RAW_DIR / "NID_National.xlsx"
    nid_data = load_nid_data(nid_path)

    # 3. Population
    pop_path = ROOT / "data" / "seed" / "county_population.csv"
    population = load_census_population(pop_path if pop_path.exists() else None)

    # 4. Compute scores
    scores = compute_reservoir_scores(nid_data, usbr_current, population)

    # 5. Save
    fieldnames = [
        "fips",
        "reservoir_capacity_af",
        "dam_count",
        "reservoir_capacity_per_capita",
        "reservoir_score",
        "usbr_storage_pct",
        "reservoir_current_score",
    ]
    out_path = OUT_DIR / "reservoirs.csv"
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for fips in sorted(scores.keys()):
            writer.writerow({"fips": fips, **scores[fips]})

    print(f"  Saved {len(scores)} counties → {out_path}")
    print("Done.\n")
    return out_path


if __name__ == "__main__":
    run()
